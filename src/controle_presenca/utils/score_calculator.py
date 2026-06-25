import os
import openpyxl
from typing import Dict, Any

class ScoreCalculator:
    _criterios = None

    @staticmethod
    def _converter_questao_id(q_val: Any) -> Any:
        try:
            return int(q_val)
        except ValueError:
            return q_val

    @staticmethod
    def _converter_pontos(pts_val: Any) -> float:
        try:
            return float(pts_val) if pts_val is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    @classmethod
    def _extrair_dados_planilha(cls, sheet) -> Dict[Any, Dict[str, float]]:
        criterios = {}
        current_q = None
        for row in range(3, sheet.max_row + 1):
            q_val = sheet.cell(row=row, column=1).value
            if q_val is not None:
                current_q = cls._converter_questao_id(q_val)
                criterios[current_q] = {}

            alt_val = sheet.cell(row=row, column=2).value
            pts_val = sheet.cell(row=row, column=3).value

            if current_q is not None and alt_val is not None:
                alt_clean = str(alt_val).strip()
                pts = cls._converter_pontos(pts_val)
                criterios[current_q][alt_clean] = pts
        return criterios

    @classmethod
    def _carregar_criterios(cls) -> Dict[int, Dict[str, float]]:
        if cls._criterios is not None:
            return cls._criterios

        # Caminho padrão para a planilha de critérios
        diretorio_atual = os.path.dirname(os.path.dirname(__file__))
        caminho_xlsx = os.path.join(diretorio_atual, 'config', 'criterios_pontuacao.xlsx')
        
        if not os.path.exists(caminho_xlsx):
            # Se não existir (ex: em ambiente de teste), retorna dicionário vazio
            return {}

        wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
        try:
            sheet = wb.active
            cls._criterios = cls._extrair_dados_planilha(sheet)
        finally:
            wb.close()
            
        return cls._criterios

    @staticmethod
    def _obter_pontuacao_resposta(criterios_questao: Dict[str, float], resposta: str) -> float:
        resp_clean = str(resposta).strip() if resposta is not None else ""
        if resp_clean in criterios_questao:
            return criterios_questao[resp_clean]
        
        # Match case-insensitive e strip-tolerant como fallback
        for key, val in criterios_questao.items():
            if key.lower() == resp_clean.lower():
                return val
        
        return 0.0

    @classmethod
    def calcular_score(cls, respostas: Dict[int, str]) -> float:
        """
        Calcula a pontuação socioeconômica total com base em um dicionário
        de respostas do candidato, mapeando {numero_questao: alternativa_escolhida}.
        """
        criterios = cls._carregar_criterios()
        score_total = 0.0

        for q_num, resposta in respostas.items():
            if q_num in criterios:
                score_total += cls._obter_pontuacao_resposta(criterios[q_num], resposta)

        return score_total
